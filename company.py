"""
company.py
특정 업체(기본: 한올바이오파마) 제품을 행으로, 각 월(스냅샷)을 열로 놓고
해당 월에 있으면 상한금액, 없으면 'X'로 채운 매트릭스를 만든다.

- 어떤 달에 사라진 품목 → 그 달 이후 열이 'X'
- 새로 등재된 품목       → 그 이전 열이 'X'
- 상한금액 변동          → 열 간 숫자가 달라짐
"""
import pandas as pd

DEFAULT_MAKER = "한올바이오파마"


def _filter_maker(df: pd.DataFrame, maker_kw: str) -> pd.DataFrame:
    return df[df["maker"].astype(str).str.contains(maker_kw, na=False)]


def cell_marks(row, months):
    """각 월 셀의 표시 상태를 반환. {month: 'x'|'new'|'up'|'down'|''}"""
    marks = {}
    prev_val = None
    seen = False
    for m in months:
        v = row[m]
        if v == "X":
            marks[m] = "x"
            continue
        if not seen:
            # 첫 등장: 첫 열이 아니면 신규 진입
            marks[m] = "new" if m != months[0] else ""
        elif prev_val is not None and float(v) != float(prev_val):
            marks[m] = "up" if float(v) > float(prev_val) else "down"
        else:
            marks[m] = ""
        prev_val = v
        seen = True
    return marks


def build_matrix(maker_kw: str = DEFAULT_MAKER, labels=None, snapshots: dict = None):
    """
    labels: 포함할 월 리스트. 미지정 시 전체.
    snapshots: {label: DataFrame} 직접 전달 시 storage 대신 사용(Streamlit/파일 기반).
    반환: (matrix_df, months)
    """
    if snapshots is not None:
        all_labels = sorted(snapshots.keys())
        loader = lambda m: snapshots[m]
    else:
        import storage
        snaps = storage.list_snapshots()
        all_labels = sorted(snaps["label"].tolist())
        loader = lambda m: storage.load_snapshot(m)
    months = [l for l in all_labels if (labels is None or l in labels)]
    if not months:
        return pd.DataFrame(), []

    # 월별 해당 업체 제품 적재
    per_month = {}
    names = {}     # code -> 가장 최근 제품명
    for m in months:
        d = _filter_maker(loader(m), maker_kw)
        per_month[m] = dict(zip(d["code"], d["price"]))
        for c, n in zip(d["code"], d["name"]):
            names[c] = n  # 뒤 월(최근)이 덮어써서 최신명 유지

    # 한올 제품으로 한 번이라도 등장한 모든 코드
    codes = sorted({c for m in months for c in per_month[m].keys()})

    rows = []
    for code in codes:
        row = {"code": code, "name": names.get(code, "")}
        present = []  # (month, price or None)
        for m in months:
            if code in per_month[m]:
                row[m] = per_month[m][code]
                present.append(per_month[m][code])
            else:
                row[m] = "X"
        # 상태 판정
        first_present = next((i for i, m in enumerate(months) if row[m] != "X"), None)
        n = len(months)
        flags = []
        # 신규: '가장 최근 달'에 처음 등장한 제품만. (이전 달에 등장한 옛 신규는 유지)
        if first_present is not None and first_present == n - 1 and n > 1:
            flags.append("신규")
        if row[months[-1]] == "X" and first_present is not None:
            flags.append("삭제")            # 최신 월에 사라짐
        prices = [p for p in present]
        if len(set(prices)) > 1:
            flags.append("변동")
        row["status"] = ", ".join(flags) if flags else "유지"
        rows.append(row)

    mat = pd.DataFrame(rows)

    # 정렬 순위: 그 해에 '변화'가 있었던 제품을 위로.
    # 삭제(0) > 신규 등장 이력(1) > 변동(2) > 유지(3).
    # 상태 라벨('신규'는 최신 달만)과 무관하게, 옛 신규도 등장 이력으로 위쪽에 배치.
    def _rank(row):
        present_vals = [row[m] for m in months if row[m] != "X"]
        first_present = next((i for i, m in enumerate(months) if row[m] != "X"), None)
        is_removed = row[months[-1]] == "X" and first_present is not None
        had_new = first_present is not None and first_present > 0   # 첫 달 이후 등장 = 신규 이력
        varied = len(set(present_vals)) > 1
        if is_removed:
            return 0
        if had_new:
            return 1
        if varied:
            return 2
        return 3

    mat["_sort"] = mat.apply(_rank, axis=1)
    mat = mat.sort_values(["_sort", "name"]).drop(columns="_sort").reset_index(drop=True)
    return mat, months


def to_excel(maker_kw: str, mat: pd.DataFrame, months, out_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook(); ws = wb.active; ws.title = "월별 상한금액"
    X_FILL = PatternFill("solid", fgColor="FFC7CE")
    NEW_FILL = PatternFill("solid", fgColor="C6EFCE")
    UP_FILL = PatternFill("solid", fgColor="FCE4D6")
    DOWN_FILL = PatternFill("solid", fgColor="DDEBF7")
    HDR = PatternFill("solid", fgColor="0D5C4A")
    HF = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E3E1D8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"{maker_kw} 제품 월별 상한금액 추적"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"기간: {months[0]} ~ {months[-1]}  |  품목 {len(mat)}개  |  "
                f"없는 달 X(빨강) · 신규(초록) · 인상(주황) · 인하(파랑)")
    ws["A2"].font = Font(size=11, color="6B7A73")

    cols = ["code", "name"] + months + ["status"]
    if "사유" in mat.columns:
        cols.append("사유")
    head_map = {"code": "제품코드", "name": "제품명", "status": "상태", "사유": "사유"}
    hr = 4
    for j, c in enumerate(cols, 1):
        cell = ws.cell(hr, j, head_map.get(c, c))
        cell.fill = HDR; cell.font = HF; cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for i, (_, r) in enumerate(mat.iterrows(), start=hr + 1):
        marks = cell_marks(r, months)
        for j, c in enumerate(cols, 1):
            v = r[c]
            cell = ws.cell(i, j, v)
            cell.border = border
            if c in months:
                cell.alignment = Alignment(horizontal="right")
                mk = marks[c]
                if mk == "x":
                    cell.fill = X_FILL; cell.font = Font(bold=True, color="B42318")
                elif mk == "new":
                    cell.fill = NEW_FILL; cell.font = Font(bold=True, color="0F7A4D")
                elif mk == "up":
                    cell.fill = UP_FILL; cell.font = Font(bold=True, color="B54708")
                elif mk == "down":
                    cell.fill = DOWN_FILL; cell.font = Font(bold=True, color="1F5FA8")
            if c == "name" and "신규" in str(r["status"]):
                cell.fill = NEW_FILL
    widths = {"code": 12, "name": 42, "status": 14, "사유": 40}
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(hr, j).column_letter].width = widths.get(c, 12)
    ws.freeze_panes = "C5"
    wb.save(out_path)
    return out_path
